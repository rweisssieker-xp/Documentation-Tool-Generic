"""
Test-Checklisten Generator: Erstellt automatisch Test-Checklisten aus dokumentierten Schritten
"""

from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
import csv

from src.utils.logger import get_logger

logger = get_logger(__name__)


class TestChecklistGenerator:
    """Generiert Test-Checklisten aus dokumentierten Schritten"""
    
    def __init__(self):
        """Initialisiert den Test-Checklist Generator"""
        pass
    
    def generate_checklist(
        self,
        steps: List[Dict],
        output_path: Path,
        format: str = 'csv',
        include_test_data: bool = False
    ) -> Path:
        """
        Generiert Test-Checkliste
        
        Args:
            steps: Liste von dokumentierten Schritten
            output_path: Ausgabepfad
            format: Format ('csv', 'excel', 'markdown')
            include_test_data: Ob Test-Daten-Spalten eingefügt werden sollen
            
        Returns:
            Pfad zur erstellten Checkliste
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if format == 'csv':
            return self._export_csv(steps, output_path, include_test_data)
        elif format == 'markdown':
            return self._export_markdown(steps, output_path, include_test_data)
        elif format == 'excel':
            return self._export_excel(steps, output_path, include_test_data)
        else:
            raise ValueError(f"Unbekanntes Format: {format}")
    
    def _export_csv(self, steps: List[Dict], output_path: Path, include_test_data: bool) -> Path:
        """Exportiert als CSV"""
        csv_path = output_path.with_suffix('.csv')
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            
            # Header
            headers = ['Schritt', 'Aktion', 'Erwartetes Ergebnis', 'Status', 'Bemerkung']
            if include_test_data:
                headers.insert(3, 'Test-Daten')
            
            writer.writerow(headers)
            
            # Schritte
            for step in steps:
                step_num = step.get('step_number', '?')
                window_title = step.get('window_title', 'Unbekannt')
                description = step.get('description', '')
                
                # Extrahiere Aktion aus Beschreibung
                action = self._extract_action(description, window_title)
                
                # Erwartetes Ergebnis
                expected_result = self._extract_expected_result(description, window_title)
                
                row = [step_num, action, expected_result, '', '']  # Status und Bemerkung leer
                
                if include_test_data:
                    row.insert(3, '')  # Test-Daten leer
                
                writer.writerow(row)
        
        logger.info(f"Test-Checkliste (CSV) exportiert: {csv_path}")
        return csv_path
    
    def _export_markdown(self, steps: List[Dict], output_path: Path, include_test_data: bool) -> Path:
        """Exportiert als Markdown"""
        md_path = output_path.with_suffix('.md')
        
        lines = [
            "# Test-Checkliste",
            "",
            f"*Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}*",
            "",
            "| Schritt | Aktion | Erwartetes Ergebnis | Status | Bemerkung |",
            "|---------|--------|---------------------|--------|-----------|"
        ]
        
        if include_test_data:
            lines[4] = "| Schritt | Aktion | Erwartetes Ergebnis | Test-Daten | Status | Bemerkung |"
            lines[5] = "|---------|--------|---------------------|------------|--------|-----------|"
        
        for step in steps:
            step_num = step.get('step_number', '?')
            window_title = step.get('window_title', 'Unbekannt')
            description = step.get('description', '')
            
            action = self._extract_action(description, window_title)
            expected_result = self._extract_expected_result(description, window_title)
            
            if include_test_data:
                row = f"| {step_num} | {action} | {expected_result} | | ⬜ | |"
            else:
                row = f"| {step_num} | {action} | {expected_result} | ⬜ | |"
            
            lines.append(row)
        
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        logger.info(f"Test-Checkliste (Markdown) exportiert: {md_path}")
        return md_path
    
    def _export_excel(self, steps: List[Dict], output_path: Path, include_test_data: bool) -> Path:
        """Exportiert als Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            excel_path = output_path.with_suffix('.xlsx')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test-Checkliste"
            
            # Header
            headers = ['Schritt', 'Aktion', 'Erwartetes Ergebnis', 'Status', 'Bemerkung']
            if include_test_data:
                headers.insert(3, 'Test-Daten')
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Schritte
            for row_num, step in enumerate(steps, start=2):
                step_num = step.get('step_number', '?')
                window_title = step.get('window_title', 'Unbekannt')
                description = step.get('description', '')
                
                action = self._extract_action(description, window_title)
                expected_result = self._extract_expected_result(description, window_title)
                
                ws.cell(row=row_num, column=1).value = step_num
                ws.cell(row=row_num, column=2).value = action
                ws.cell(row=row_num, column=3).value = expected_result
                
                if include_test_data:
                    ws.cell(row=row_num, column=4).value = ''  # Test-Daten
                    ws.cell(row=row_num, column=5).value = ''  # Status
                    ws.cell(row=row_num, column=6).value = ''  # Bemerkung
                else:
                    ws.cell(row=row_num, column=4).value = ''  # Status
                    ws.cell(row=row_num, column=5).value = ''  # Bemerkung
            
            # Auto-Breite
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            
            logger.info(f"Test-Checkliste (Excel) exportiert: {excel_path}")
            return excel_path
        
        except ImportError:
            logger.warning("openpyxl nicht verfügbar, verwende CSV-Export")
            return self._export_csv(steps, output_path, include_test_data)
    
    def _extract_action(self, description: str, window_title: str) -> str:
        """Extrahiert Aktion aus Beschreibung"""
        if not description:
            return window_title
        
        # Entferne Präfixe
        desc = description.strip()
        if ':' in desc:
            parts = desc.split(':', 1)
            if len(parts) > 1:
                desc = parts[1].strip()
        
        # Kürze auf erste Satz/Aktion
        sentences = desc.split('.')
        if sentences:
            action = sentences[0].strip()
            if len(action) > 100:
                action = action[:97] + "..."
            return action
        
        return desc[:100] if len(desc) > 100 else desc
    
    def _extract_expected_result(self, description: str, window_title: str) -> str:
        """Extrahiert erwartetes Ergebnis aus Beschreibung"""
        if not description:
            return f"Fenster '{window_title}' wird angezeigt"
        
        # Suche nach Ergebnis-Indikatoren
        desc_lower = description.lower()
        
        # Erkenne Ergebnis-Phrasen
        result_phrases = [
            'erscheint', 'wird angezeigt', 'wird geöffnet', 'wird erstellt',
            'ist sichtbar', 'wird gespeichert', 'wird aktualisiert'
        ]
        
        for phrase in result_phrases:
            if phrase in desc_lower:
                # Extrahiere Satz mit Phrase
                sentences = description.split('.')
                for sentence in sentences:
                    if phrase in sentence.lower():
                        return sentence.strip()
        
        # Fallback: Letzter Satz
        sentences = description.split('.')
        if len(sentences) > 1:
            return sentences[-1].strip()
        
        return "Erfolgreich abgeschlossen"

